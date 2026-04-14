"""
DataForge - Excel Engine
=========================
Advanced Excel read/write operations with formatting,
multi-sheet support, and styled output reports.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    PatternFill,
    Side,
    numbers,
)
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference
from loguru import logger

from config import OUTPUT_DIR


class ExcelEngine:
    """
    Production Excel engine for reading messy files and
    producing beautifully formatted output workbooks.
    """

    # ── Reading ──────────────────────────────────────────────

    @staticmethod
    def read(
        filepath: str | Path,
        sheet_name: str | int | None = None,
        header_row: int = 0,
        skip_rows: list[int] | None = None,
        use_cols: list[str] | None = None,
    ) -> pd.DataFrame | dict[str, pd.DataFrame]:
        """
        Read Excel/CSV files with smart detection.

        Returns a single DataFrame or dict of DataFrames (multi-sheet).
        """
        filepath = Path(filepath)
        if not filepath.exists():
            raise FileNotFoundError(f"File not found: {filepath}")

        logger.info(f"Reading file: {filepath.name}")

        if filepath.suffix.lower() == ".csv":
            # Try different encodings
            for encoding in ("utf-8", "latin-1", "cp1252"):
                try:
                    return pd.read_csv(
                        filepath,
                        header=header_row,
                        skiprows=skip_rows,
                        usecols=use_cols,
                        encoding=encoding,
                    )
                except UnicodeDecodeError:
                    continue
            raise ValueError(f"Cannot decode CSV file: {filepath}")

        # Excel file
        kwargs: dict[str, Any] = {
            "header": header_row,
            "skiprows": skip_rows,
            "usecols": use_cols,
            "engine": "openpyxl",
        }

        if sheet_name is not None:
            kwargs["sheet_name"] = sheet_name

        result = pd.read_excel(filepath, **kwargs)

        if isinstance(result, dict):
            total_rows = sum(len(df) for df in result.values())
            logger.info(f"Loaded {len(result)} sheets, {total_rows} total rows")
        else:
            logger.info(f"Loaded {len(result)} rows, {len(result.columns)} columns")

        return result

    # ── Writing ──────────────────────────────────────────────

    @staticmethod
    def write(
        data: pd.DataFrame | dict[str, pd.DataFrame],
        filename: str,
        output_dir: Path | None = None,
        styled: bool = True,
        include_charts: bool = False,
        chart_config: dict | None = None,
    ) -> Path:
        """
        Write DataFrame(s) to a beautifully formatted Excel file.

        Args:
            data: Single DataFrame or dict of {sheet_name: DataFrame}
            filename: Output filename (without extension)
            styled: Apply professional formatting
            include_charts: Auto-generate charts for numeric data
        """
        out_dir = output_dir or OUTPUT_DIR
        out_dir.mkdir(parents=True, exist_ok=True)
        filepath = out_dir / f"{filename}.xlsx"

        # Normalize input
        if isinstance(data, pd.DataFrame):
            sheets = {"Sheet1": data}
        else:
            sheets = data

        wb = Workbook()
        wb.remove(wb.active)  # Remove default sheet

        for sheet_name, df in sheets.items():
            ws = wb.create_sheet(title=sheet_name[:31])  # Excel 31-char limit
            ExcelEngine._write_dataframe_to_sheet(ws, df)

            if styled:
                ExcelEngine._apply_styling(ws, df)

            if include_charts:
                ExcelEngine._add_auto_chart(ws, df, chart_config)

        wb.save(filepath)
        logger.info(f"Saved Excel file: {filepath}")
        return filepath

    # ── Private Helpers ──────────────────────────────────────

    @staticmethod
    def _write_dataframe_to_sheet(ws, df: pd.DataFrame) -> None:
        """Write a DataFrame to an openpyxl worksheet."""
        # Write headers
        for col_idx, col_name in enumerate(df.columns, 1):
            ws.cell(row=1, column=col_idx, value=str(col_name))

        # Write data
        for row_idx, row in enumerate(df.itertuples(index=False), 2):
            for col_idx, value in enumerate(row, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                # Handle NaN
                if pd.isna(value):
                    cell.value = None
                else:
                    cell.value = value

    @staticmethod
    def _apply_styling(ws, df: pd.DataFrame) -> None:
        """Apply professional formatting to worksheet."""
        # Color palette
        header_fill = PatternFill(start_color="1B2A4A", end_color="1B2A4A", fill_type="solid")
        header_font = Font(name="Calibri", bold=True, color="FFFFFF", size=11)
        alt_fill = PatternFill(start_color="F0F4F8", end_color="F0F4F8", fill_type="solid")
        border = Border(
            left=Side(style="thin", color="D1D5DB"),
            right=Side(style="thin", color="D1D5DB"),
            top=Side(style="thin", color="D1D5DB"),
            bottom=Side(style="thin", color="D1D5DB"),
        )

        num_cols = len(df.columns)
        num_rows = len(df) + 1  # +1 for header

        # Style headers
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = border

        # Style data rows
        for row in range(2, num_rows + 1):
            for col in range(1, num_cols + 1):
                cell = ws.cell(row=row, column=col)
                cell.border = border
                cell.alignment = Alignment(vertical="center")
                cell.font = Font(name="Calibri", size=10)

                # Alternating row colors
                if row % 2 == 0:
                    cell.fill = alt_fill

        # Auto-fit column widths
        for col_idx in range(1, num_cols + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for row_idx in range(1, num_rows + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            adjusted_width = min(max_length + 4, 50)
            ws.column_dimensions[column_letter].width = adjusted_width

        # Freeze top row
        ws.freeze_panes = "A2"

        # Add auto-filter
        ws.auto_filter.ref = f"A1:{get_column_letter(num_cols)}{num_rows}"

    @staticmethod
    def _add_auto_chart(
        ws, df: pd.DataFrame, config: dict | None = None
    ) -> None:
        """Auto-generate a bar chart for the first numeric column."""
        numeric_cols = df.select_dtypes(include=["number"]).columns.tolist()
        if not numeric_cols:
            return

        chart_col = numeric_cols[0]
        col_idx = df.columns.tolist().index(chart_col) + 1
        num_rows = len(df) + 1

        chart = BarChart()
        chart.type = "col"
        chart.title = f"{chart_col} Overview"
        chart.style = 10
        chart.y_axis.title = chart_col
        chart.x_axis.title = df.columns[0] if len(df.columns) > 1 else "Index"

        data = Reference(ws, min_col=col_idx, min_row=1, max_row=num_rows)
        categories = Reference(ws, min_col=1, min_row=2, max_row=num_rows)
        chart.add_data(data, titles_from_data=True)
        chart.set_categories(categories)
        chart.shape = 4

        # Place chart to the right of data
        chart_position = f"{get_column_letter(len(df.columns) + 2)}2"
        ws.add_chart(chart, chart_position)
        logger.debug(f"Added chart for column '{chart_col}'")
